#!/usr/bin/env python3
"""
Script para generar template de Excel para importación de clientes y equipos
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
import os

def crear_template_importacion():
    """
    Crea un archivo Excel con el template para importar clientes y equipos
    """
    
    # Crear DataFrame con las columnas necesarias
    columnas = [
        # === INFORMACIÓN DEL CLIENTE ===
        'TIPO_CLIENTE',  # EMPRESA, PARTICULAR, ORGANISMO_PUBLICO
        'SUCURSAL',  # Nombre de la sucursal
        'RAZON_SOCIAL',  # Nombre legal de la empresa
        'NOMBRE_FANTASIA',  # Nombre comercial (opcional)
        'CUIT',  # CUIT del cliente
        'EMAIL',  # Email del cliente
        'TELEFONO',  # Teléfono del cliente
        'DIRECCION',  # Dirección completa
        'CODIGO_POSTAL',  # Código postal
        'CIUDAD',  # Nombre de la ciudad
        'PROVINCIA',  # Nombre de la provincia
        'OBSERVACIONES_CLIENTE',  # Observaciones del cliente
        
        # === INFORMACIÓN DE CONTACTO ===
        'NOMBRE_CONTACTO',  # Nombre del contacto principal
        'APELLIDO_CONTACTO',  # Apellido del contacto principal
        'ROL_CONTACTO',  # GERENTE, JEFE_TALLER, ADMINISTRATIVO, COMPRAS, OTRO
        'EMAIL_CONTACTO',  # Email del contacto
        'TELEFONO_FIJO_CONTACTO',  # Teléfono fijo del contacto
        'TELEFONO_CELULAR_CONTACTO',  # Teléfono celular del contacto
        'ES_CONTACTO_PRINCIPAL',  # TRUE/FALSE
        
        # === INFORMACIÓN DEL EQUIPO ===
        'TIPO_EQUIPO',  # Nombre del tipo de equipo (ej: Retroexcavadora)
        'MODELO_EQUIPO',  # Nombre del modelo (ej: 310SL)
        'MARCA_EQUIPO',  # Marca del equipo (ej: John Deere)
        'NUMERO_SERIE',  # Número de serie/PIN del equipo
        'MODELO_MOTOR',  # Modelo del motor (opcional)
        'NUMERO_SERIE_MOTOR',  # Número de serie del motor (opcional)
        'AÑO_FABRICACION',  # Año de fabricación del equipo
        'FECHA_VENTA',  # Fecha de venta (YYYY-MM-DD)
        'NOTAS_EQUIPO',  # Notas sobre el equipo
        'HOROMETRO_ACTUAL',  # Horas actuales del equipo
    ]
    
    # Crear datos de ejemplo
    datos_ejemplo = [
        # Cliente 1 - Empresa con equipo
        ['EMPRESA', 'RGA', 'Constructora Patagonia S.A.', 'Patagonia Construcciones', 
         '20-12345678-9', 'info@patagonia.com', '0299-1234567', 'Av. San Martín 1234', 
         '8300', 'Neuquén', 'Neuquén', 'Cliente importante del sector construcción',
         'Juan', 'Pérez', 'GERENTE', 'juan.perez@patagonia.com', '0299-1234567', 
         '299-1234567', 'TRUE',
         'Retroexcavadora', '310SL', 'John Deere', 'JD123456789', 'PowerTech 4045T', 
         'MOT123456', '2020', '2020-03-15', 'Equipo en excelente estado', '1250.50'],
        
        # Cliente 2 - Particular con equipo
        ['PARTICULAR', 'RGA', 'María González', '', 
         '27-98765432-1', 'maria.gonzalez@gmail.com', '0299-9876543', 'Calle Mitre 567', 
         '8300', 'Neuquén', 'Neuquén', 'Cliente particular',
         'María', 'González', 'OTRO', 'maria.gonzalez@gmail.com', '', 
         '299-9876543', 'TRUE',
         'Motoniveladora', '670G', 'John Deere', 'JD987654321', 'PowerTech 6068T', 
         'MOT987654', '2019', '2019-08-20', 'Equipo para uso personal', '890.25'],
        
        # Cliente 3 - Organismo público con equipo
        ['ORGANISMO_PUBLICO', 'RGA', 'Municipalidad de Neuquén', 'Municipalidad', 
         '30-12345678-9', 'info@muni-neuquen.gov.ar', '0299-4567890', 'Av. Argentina 1000', 
         '8300', 'Neuquén', 'Neuquén', 'Organismo público municipal',
         'Carlos', 'Rodríguez', 'JEFE_TALLER', 'carlos.rodriguez@muni-neuquen.gov.ar', 
         '0299-4567890', '299-4567890', 'TRUE',
         'Grupo Electrógeno', 'PP100', 'PowerPro', 'PP100123456', 'Cummins 6BTA5.9', 
         'CUM123456', '2021', '2021-01-10', 'Equipo para emergencias', '150.75'],
    ]
    
    # Crear DataFrame
    df = pd.DataFrame(datos_ejemplo, columns=columnas)
    
    # Crear archivo Excel
    nombre_archivo = f'template_importacion_clientes_equipos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Template', index=False)
        
        # Obtener el workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets['Template']
        
        # Aplicar estilos
        aplicar_estilos_excel(workbook, worksheet, df)
        
        # Crear hoja de instrucciones
        crear_hoja_instrucciones(workbook)
        
        # Crear hoja de catálogos
        crear_hoja_catalogos(workbook)
    
    print(f"✅ Template creado exitosamente: {nombre_archivo}")
    print(f"📁 Ubicación: {os.path.abspath(nombre_archivo)}")
    
    return nombre_archivo

def aplicar_estilos_excel(workbook, worksheet, df):
    """Aplica estilos al Excel"""
    
    # Definir colores
    color_encabezado = "366092"
    color_cliente = "E7E6F7"
    color_contacto = "FCD5B4"
    color_equipo = "D5E8D4"
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=color_encabezado, end_color=color_encabezado, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar estilos a encabezados
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Aplicar colores de fondo por sección
    # Sección Cliente (columnas 1-12)
    for col in range(1, 13):
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=color_cliente, end_color=color_cliente, fill_type="solid")
    
    # Sección Contacto (columnas 13-19)
    for col in range(13, 20):
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=color_contacto, end_color=color_contacto, fill_type="solid")
    
    # Sección Equipo (columnas 20-29)
    for col in range(20, 30):
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=color_equipo, end_color=color_equipo, fill_type="solid")
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def crear_hoja_instrucciones(workbook):
    """Crea hoja con instrucciones de uso"""
    
    worksheet = workbook.create_sheet("INSTRUCCIONES")
    
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTACIÓN DE CLIENTES Y EQUIPOS"],
        [""],
        ["1. INFORMACIÓN DEL CLIENTE (Columnas A-L):"],
        ["   • TIPO_CLIENTE: EMPRESA, PARTICULAR, ORGANISMO_PUBLICO"],
        ["   • SUCURSAL: Nombre exacto de la sucursal (RGA, RIO, etc.)"],
        ["   • RAZON_SOCIAL: Nombre legal de la empresa"],
        ["   • NOMBRE_FANTASIA: Nombre comercial (opcional)"],
        ["   • CUIT: Formato XX-XXXXXXXX-X"],
        ["   • EMAIL: Email válido del cliente"],
        ["   • TELEFONO: Teléfono del cliente"],
        ["   • DIRECCION: Dirección completa"],
        ["   • CODIGO_POSTAL: Código postal"],
        ["   • CIUDAD: Nombre de la ciudad"],
        ["   • PROVINCIA: Nombre de la provincia"],
        ["   • OBSERVACIONES_CLIENTE: Notas sobre el cliente"],
        [""],
        ["2. INFORMACIÓN DE CONTACTO (Columnas M-S):"],
        ["   • NOMBRE_CONTACTO: Nombre del contacto principal"],
        ["   • APELLIDO_CONTACTO: Apellido del contacto principal"],
        ["   • ROL_CONTACTO: GERENTE, JEFE_TALLER, ADMINISTRATIVO, COMPRAS, OTRO"],
        ["   • EMAIL_CONTACTO: Email del contacto"],
        ["   • TELEFONO_FIJO_CONTACTO: Teléfono fijo (opcional)"],
        ["   • TELEFONO_CELULAR_CONTACTO: Teléfono celular (opcional)"],
        ["   • ES_CONTACTO_PRINCIPAL: TRUE o FALSE"],
        [""],
        ["3. INFORMACIÓN DEL EQUIPO (Columnas T-AC):"],
        ["   • TIPO_EQUIPO: Nombre del tipo (ej: Retroexcavadora, Motoniveladora)"],
        ["   • MODELO_EQUIPO: Nombre del modelo (ej: 310SL, 670G)"],
        ["   • MARCA_EQUIPO: Marca del equipo (ej: John Deere, PowerPro)"],
        ["   • NUMERO_SERIE: Número de serie/PIN único"],
        ["   • MODELO_MOTOR: Modelo del motor (opcional)"],
        ["   • NUMERO_SERIE_MOTOR: Número de serie del motor (opcional)"],
        ["   • AÑO_FABRICACION: Año de fabricación (YYYY)"],
        ["   • FECHA_VENTA: Fecha de venta (YYYY-MM-DD)"],
        ["   • NOTAS_EQUIPO: Notas sobre el equipo"],
        ["   • HOROMETRO_ACTUAL: Horas actuales del equipo"],
        [""],
        ["4. REGLAS IMPORTANTES:"],
        ["   • No modificar los nombres de las columnas"],
        ["   • Mantener el formato de fecha YYYY-MM-DD"],
        ["   • Los campos obligatorios no pueden estar vacíos"],
        ["   • El CUIT debe ser único en el sistema"],
        ["   • El número de serie del equipo debe ser único"],
        ["   • Verificar que sucursal, ciudad y provincia existan en el sistema"],
        [""],
        ["5. PROCESO DE IMPORTACIÓN:"],
        ["   • Completar los datos en la hoja 'Template'"],
        ["   • Guardar el archivo"],
        ["   • Subir el archivo en la sección de importación"],
        ["   • Revisar los resultados de la importación"],
    ]
    
    for i, instruccion in enumerate(instrucciones, 1):
        worksheet.cell(row=i, column=1, value=instruccion[0])
    
    # Ajustar ancho de columna
    worksheet.column_dimensions['A'].width = 80

def crear_hoja_catalogos(workbook):
    """Crea hoja con catálogos de valores válidos"""
    
    worksheet = workbook.create_sheet("CATÁLOGOS")
    
    catalogos = [
        ["CATÁLOGOS DE VALORES VÁLIDOS"],
        [""],
        ["TIPO_CLIENTE:"],
        ["• EMPRESA"],
        ["• PARTICULAR"],
        ["• ORGANISMO_PUBLICO"],
        [""],
        ["SUCURSAL:"],
        ["• RGA"],
        ["• RIO"],
        ["• Otras sucursales según el sistema"],
        [""],
        ["ROL_CONTACTO:"],
        ["• GERENTE"],
        ["• JEFE_TALLER"],
        ["• ADMINISTRATIVO"],
        ["• COMPRAS"],
        ["• OTRO"],
        [""],
        ["TIPO_EQUIPO (Ejemplos):"],
        ["• Retroexcavadora"],
        ["• Motoniveladora"],
        ["• Grupo Electrógeno"],
        ["• Excavadora"],
        ["• Cargador Frontal"],
        ["• Otros tipos según el sistema"],
        [""],
        ["MARCA_EQUIPO (Ejemplos):"],
        ["• John Deere"],
        ["• PowerPro"],
        ["• Caterpillar"],
        ["• Komatsu"],
        ["• Otras marcas según el sistema"],
        [""],
        ["FORMATO DE FECHAS:"],
        ["• YYYY-MM-DD (ej: 2023-12-25)"],
        [""],
        ["FORMATO DE CUIT:"],
        ["• XX-XXXXXXXX-X (ej: 20-12345678-9)"],
    ]
    
    for i, catalogo in enumerate(catalogos, 1):
        worksheet.cell(row=i, column=1, value=catalogo[0])
    
    # Ajustar ancho de columna
    worksheet.column_dimensions['A'].width = 50

if __name__ == "__main__":
    crear_template_importacion() 